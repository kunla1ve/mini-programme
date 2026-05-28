#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
交易記錄處理工具 - 打包版
支持 PyInstaller 打包
"""

import re
import os
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import copy


def get_resource_path(relative_path):
    """获取资源的绝对路径，支持开发环境和打包后的exe"""
    try:
        # PyInstaller 创建临时文件夹，将路径存储在 _MEIPASS 中
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)


# ============================================================================
# 交易記錄文本解析
# ============================================================================

def parse_trading_data(text):
    """解析交易记录文本"""
    
    # 清理文本
    text = re.sub(r' -\s*\n\s*\n\s*(POEMSHK|MCSGXPHL\w*|PHLX\w*)', r' - \1', text)
    text = re.sub(r'(@\s*[\d.]+\s*)\n\s*\n\s*(POEMSHK|MCSGXPHL\w*|PHLX\w*)', 
                  r'\1- \1', text)
    
    lines = text.split('\n')
    
    patterns = [
        r'(買入|賣出|BUY|SELL)\s+(\d+)\s*(?:\[(\d+)\])?\s*([A-Za-z0-9.$]+?)\s*@\s*([\d.]+)\s*-',
        r'(買入|賣出|BUY|SELL)\s+(\d+)\s+([A-Za-z0-9.$]+?)\s*@\s*([\d.]+)',
    ]
    
    transactions = []
    current_symbol = None
    current_action = None
    current_details = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        if (line.startswith('––––––') or line.startswith('---') or
            re.search(r'^[A-Za-z]+:\s*(January|February|March|April|May|June|July|August|September|October|November|December|Call|Put|Option)', line) or
            re.search(r'^(POEMSHK|MCSGXPHL|PHLX)\s*\(', line) or
            re.match(r'^\d{2}/\d{2}/\d{4}$', line) or
            re.search(r'^[A-Za-z\s/()-]+:\s*(January|February|March|April|May|June|July|August|September|October|November|December|20\d{2})', line)):
            continue
        
        for pattern in patterns:
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                action_text = match.group(1).upper()
                qty = int(match.group(3)) if match.group(3) and pattern == patterns[0] else int(match.group(2))
                symbol = match.group(4).strip() if pattern == patterns[0] else match.group(3).strip()
                price = float(match.group(5)) if pattern == patterns[0] else float(match.group(4))
                
                action = '沽出' if action_text in ['賣出', 'SELL'] else '買入'
                
                if current_symbol and (symbol != current_symbol or action != current_action):
                    if current_details:
                        total_qty = sum(q for q, _ in current_details)
                        total_value = sum(q * p for q, p in current_details)
                        avg_price = total_value / total_qty if total_qty > 0 else 0
                        transactions.append({
                            'action': current_action,
                            'product': current_symbol,
                            'quantity': total_qty,
                            'avg_price': avg_price
                        })
                    current_details = []
                
                current_symbol = symbol
                current_action = action
                current_details.append((qty, price))
                break
    
    if current_symbol and current_details:
        total_qty = sum(q for q, _ in current_details)
        total_value = sum(q * p for q, p in current_details)
        avg_price = total_value / total_qty if total_qty > 0 else 0
        transactions.append({
            'action': current_action,
            'product': current_symbol,
            'quantity': total_qty,
            'avg_price': avg_price
        })
    
    # 合併相同產品和方向
    merged = {}
    for trans in transactions:
        key = (trans['action'], trans['product'])
        if key in merged:
            old_qty = merged[key]['quantity']
            old_value = merged[key]['avg_price'] * old_qty
            new_qty = old_qty + trans['quantity']
            new_value = old_value + (trans['avg_price'] * trans['quantity'])
            merged[key]['quantity'] = new_qty
            merged[key]['avg_price'] = new_value / new_qty if new_qty > 0 else 0
        else:
            merged[key] = trans.copy()
    
    return list(merged.values())


# ============================================================================
# XLSX模板填充
# ============================================================================

def fill_xlsx_template(template_path, output_path, transaction_date, transactions):
    """填充數據，保留圖片，日期粗體"""
    
    sell_trans = [t for t in transactions if t['action'] == '沽出']
    buy_trans = [t for t in transactions if t['action'] == '買入']
    
    print(f"\n📊 交易統計:")
    print(f"  沽出: {len(sell_trans)} 筆")
    for t in sell_trans:
        print(f"    - {t['product']}: {t['quantity']}張 @ {t['avg_price']:.4f}")
    print(f"  買入: {len(buy_trans)} 筆")
    for t in buy_trans:
        print(f"    - {t['product']}: {t['quantity']}張 @ {t['avg_price']:.4f}")
    
    # 加載模板（保留圖片）
    wb = load_workbook(template_path)
    ws = wb.active
    
    print(f"\n📋 模板加載完成，圖片數量: {len(ws._images)}")
    
    # 定義樣式
    header_font = Font(name='Times New Roman', size=11, bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Times New Roman', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. 更新交易日期 - 在模板中搜索並更新
    date_updated = False
    for row in ws.iter_rows(min_row=1, max_row=min(8, ws.max_row), max_col=ws.max_column):
        for cell in row:
            if cell.value is not None:
                cell_str = str(cell.value)
                # 查找包含日期格式的內容
                if re.search(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', cell_str):
                    # 替換日期時間
                    new_value = re.sub(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}\s+\d{2}:\d{2}', 
                                      transaction_date, cell_str)
                    cell.value = new_value
                    cell.font = Font(name='Times New Roman', size=11, bold=True)
                    date_updated = True
                    print(f"✅ 日期已更新: {transaction_date} (行{cell.row}, 列{cell.column})")
                    break
        if date_updated:
            break
    
    if not date_updated:
        print("⚠️ 未找到日期單元格，請檢查模板格式")
    
    # 2. 清除第9行及以後的內容，但保留合併單元格結構
    max_row_to_clear = max(ws.max_row, 50)
    
    # 先記錄第9行及以後的合併單元格
    merges_to_remove = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= 9:
            merges_to_remove.append(str(merged_range))
    
    # 取消第9行及以後的合併單元格
    for merge_str in merges_to_remove:
        try:
            ws.unmerge_cells(merge_str)
        except:
            pass
    
    # 清除數據
    for row in ws.iter_rows(min_row=9, max_row=max_row_to_clear, max_col=ws.max_column):
        for cell in row:
            try:
                cell.value = None
            except:
                pass
    
    # 3. 從第9行開始寫入沽出數據
    current_row = 9
    
    # 沽出標題
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    sell_title_cell = ws.cell(row=current_row, column=1, value='沽出')
    sell_title_cell.font = Font(name='Times New Roman', size=12, bold=True)
    sell_title_cell.alignment = Alignment(horizontal='center', vertical='center')
    sell_title_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.border = thin_border
        cell.font = Font(name='Times New Roman', size=12, bold=True)
    current_row += 1
    
    # 沽出表頭
    sell_headers = ['產品', '數量', '價格', '類型']
    for col, header in enumerate(sell_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    current_row += 1
    
    # 沽出數據
    for trans in sell_trans:
        values = [
            trans['product'],
            trans['quantity'],
            round(trans['avg_price'], 4),
            '平均價'
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
        current_row += 1
    
    # 空一行
    current_row += 1
    
    # 4. 寫入買入數據
    # 買入標題
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    buy_title_cell = ws.cell(row=current_row, column=1, value='買入')
    buy_title_cell.font = Font(name='Times New Roman', size=12, bold=True)
    buy_title_cell.alignment = Alignment(horizontal='center', vertical='center')
    buy_title_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.border = thin_border
        cell.font = Font(name='Times New Roman', size=12, bold=True)
    current_row += 1
    
    # 買入表頭
    buy_headers = ['產品', '數量', '價格', '類型']
    for col, header in enumerate(buy_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    current_row += 1
    
    # 買入數據
    for trans in buy_trans:
        values = [
            trans['product'],
            trans['quantity'],
            round(trans['avg_price'], 4),
            '平均價'
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
        current_row += 1
    
    # 調整列寬
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    
    # 保存文件
    wb.save(output_path)
    wb.close()
    
    print(f"\n✅ 文件已保存: {output_path}")
    
    return True


# ============================================================================
# 主程序
# ============================================================================

def main():
    """主程序"""
    print("=" * 60)
    print("📈 交易記錄處理工具")
    print("=" * 60)
    
    print("\n請粘貼交易數據（連續按兩次回車結束輸入）：")
    
    lines = []
    blank_count = 0
    while True:
        try:
            line = input()
            if line == "":
                blank_count += 1
                if blank_count >= 2:
                    break
            else:
                blank_count = 0
            lines.append(line)
        except EOFError:
            break
    
    text = '\n'.join(lines)
    
    if not text.strip():
        print("❌ 未輸入任何數據")
        input("\n按回車鍵退出...")
        return
    
    transactions = parse_trading_data(text)
    
    if not transactions:
        print("❌ 未找到有效的交易記錄")
        input("\n按回車鍵退出...")
        return
    
    print(f"\n✅ 解析到 {len(transactions)} 筆交易:")
    for i, trans in enumerate(transactions, 1):
        print(f"  {i:2d}. {trans['action']}: {trans['product']:25s} "
              f"{trans['quantity']:4d}張 @ {trans['avg_price']:12.4f}")
    
    transaction_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # 使用打包后的资源路径
    template_path = get_resource_path("format.xlsx")
    
    if not os.path.exists(template_path):
        print(f"\n❌ 模板文件不存在: {template_path}")
        input("\n按回車鍵退出...")
        return
    
    try:
        desktop = Path.home() / "Desktop"
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        xlsx_filename = f"交易单_{timestamp}.xlsx"
        xlsx_path = os.path.join(str(desktop), xlsx_filename)
        
        fill_xlsx_template(template_path, xlsx_path, transaction_date, transactions)
        
        print(f"\n✨ 完成！")
        print(f"📁 {xlsx_filename}")
        
        open_file = input("\n是否打開文件？(y/n): ").strip().lower()
        if open_file == 'y':
            os.startfile(xlsx_path)
        
    except Exception as e:
        print(f"\n❌ 錯誤: {e}")
        import traceback
        traceback.print_exc()
    
    input("\n按回車鍵退出...")


if __name__ == "__main__":
    main()